from core import tools
import openpyxl
import os
import time


# 根据table_1.csv和students_list.csv筛选出未完成名单
def study_records_arrangement(original_date_path='.\\Original Study Records',
                              original_date_name='table_1.csv') -> list:
    """通过青年大学习已完成名单导出未完成名单和支部完成率
        导出文件的命名规则为[青年大学习学习记录](原始数据导出时间)

    :param original_date_path:
    :param original_date_name:
    :return: [unfinished_list, finished_rate_list]
        unfinished_list: 未完成名单，是一个列表
        finished_rate_list：支部完成率，是一个列表
        out_date：原始数据文件修改时间

    """
    # 加载学生名单
    students_list_name = r'students_list.csv'  # 学生名单文件名（csv格式）
    students_list = list(tools.read_csv(students_list_name))  # 列表元素是一个包含姓名、学号、支部的列表
    del students_list[0]  # 删除表头，排除干扰
    for i in students_list:
        print(i)
    input()
    # 加载已学习名单
    original_study_record_name = os.path.join(original_date_path, original_date_name)
    original_study_record = list(tools.read_csv(original_study_record_name))

    # 创建未完成名单
    unfinished_title = ('姓名', '学号', '支部代号')
    unfinished_list = []

    # 创建支部完成率
    finished_title = ('排名', '支部', '未完成人数', '已完成人数', '总人数', '完成率')
    finished_rate = {}  # 键为支部代号，元素为一个列表，列表0，1，2，3分别是未完成人数、已完成人数、总人数、完成率

    # 判断学生学生姓名或学号是否在已学习名单内
    for student in students_list:  # student的0，1，2分别代表姓名、学号、支部
        if len(student) == 0:
            continue
        # 计算支部人数
        class_code = student[2]
        if class_code in finished_rate:
            finished_rate[class_code][1] += 1  # 计算支部已完成人数（默认已完成）
            finished_rate[class_code][2] += 1  # 计算支部人数
            finished_rate[class_code][3] = finished_rate[class_code][1] / finished_rate[class_code][2]  # 计算支部完成率
        else:
            finished_rate[class_code] = [0, 1, 1, 1]
        if not tools.finished(student, original_study_record):
            unfinished_list.append(student)
            finished_rate[class_code][1] -= 1  # 计算支部已完成人数（减去没有完成的人）
            finished_rate[class_code][0] += 1  # 计算支部未完成人数
            finished_rate[class_code][3] = finished_rate[class_code][1] / finished_rate[class_code][2]  # 计算支部完成率

    # 整合支部未完成名单数据
    unfinished_list = [unfinished_title] + unfinished_list

    # 按照支部完成率对班级进行排序
    finished_rate = sorted(finished_rate.items(), key=lambda elem: elem[1][3], reverse=True)

    # 整合支部完成率数据
    finished_rate_list = [finished_title]
    rank = 1
    for i in finished_rate:
        t = [rank, i[0]] + i[1]
        t[-1] = f"{t[-1]:.2%}"
        finished_rate_list.append(t)
        rank += 1

    # 创建最终需要导出的表格
    study_record = openpyxl.Workbook()
    study_record.create_sheet('未完成名单')
    study_record.create_sheet('支部完成率')
    del study_record['Sheet']

    # 将未完成名单写入表格
    unfinished_sheet = study_record['未完成名单']
    unfinished_sheet = tools.list_to_sheet(unfinished_list, unfinished_sheet)

    # 设置未完成名单格式
    unfinished_sheet.column_dimensions['A'].width = 9.78
    unfinished_sheet.column_dimensions['B'].width = 12.56
    unfinished_sheet.column_dimensions['C'].width = 9.78

    # 将支部完成率写入表格
    finished_rate_sheet = study_record['支部完成率']
    finished_rate_sheet = tools.list_to_sheet(finished_rate_list, finished_rate_sheet)

    # 设置支部完成率格式
    finished_rate_sheet.column_dimensions['A'].width = 8.11
    finished_rate_sheet.column_dimensions['B'].width = 8.11
    finished_rate_sheet.column_dimensions['C'].width = 10.56
    finished_rate_sheet.column_dimensions['D'].width = 10.56

    # TODO 记录原始表格导出日期
    # FIXME 使用文件修改时间作为导出日期，在恶意修改文件以及其他情况下，可能会导致bug
    out_date = tools.get_out_date(original_study_record_name)

    # TODO 导出表格，根据导出时间按周归类（未完成）
    file_name = f'[青年大学习学习记录]{out_date}.xlsx'
    tools.save_excel(study_record, file_name, save_dir='.\\Study Records')

    return [unfinished_list, finished_rate_list, out_date]

